import string
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import get_column_letter


def generate(titles: list, fullname=True, **sheets: list[list[str]]):
    wb = Workbook()
    del wb["Sheet"]
    if fullname:
        titles = [" ".join(titles[:3])] + list(titles[3:])
    for sheet, cells in sheets.items():
        width: list[int] = []  # Список для получения максимальной длины строки в каждом столбце
        ws: Worksheet = wb.create_sheet(sheet)
        for num, title in enumerate(titles, 1):
            width.append(len(title))

            cell: Cell = ws.cell(1, num, title)
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style="medium"),
                right=Side(style="medium"),
                top=Side(style="medium"),
                bottom=Side(style="medium"),
            )
            ws.column_dimensions[
                string.ascii_uppercase[num - 1]
            ].width = len(title) + 2

        for row_num, row in enumerate(cells, 2):
            if fullname:
                row = [" ".join(row[:3])] + list(row[3:])
            for col_num, val in enumerate(row, 1):
                width[col_num - 1] = max(width[col_num - 1], len(str(val)))
                cell: Cell = ws.cell(row_num, col_num, val)
            for col_num in range(1, len(titles) + 1):
                cell: Cell = ws.cell(row_num, col_num)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="medium"),
                    bottom=Side(style="thin"),
                )
        for col, w in enumerate(width, 1):
            letter = get_column_letter(col)
            column = ws.column_dimensions[letter]
            column.width = max(w + 2, 4)
        # ws.row_dimensions[1].height = 150
    gen_id = datetime.now().strftime("%d-%m-%y___%H-%M-%S--%f")
    wb.save(f"app/documents/generated/xlsx/table{gen_id}.xlsx")
    return f"documents/generated/xlsx/table{gen_id}.xlsx"


if __name__ == '__main__':
    generate(
        [
            f"Заголовок{t}" for t in range(1, 6)
        ], лист1=[[
            f"Столбец{t}" for t in range(1, 6)
        ], [
            f"Столбец{t}" for t in range(1, 6)
        ]], лист2=[[
            f"Столбец{t}" for t in range(1, 6)
        ], [
            f"Столбец{t}" for t in range(1, 6)
        ]], лист3=[["1" for n in range(5)]] + [[
            f"={t}{n} * {num}" for num, t in enumerate("ABCDE", 2)
            ] for n in range(2, 17)
        ]
    )